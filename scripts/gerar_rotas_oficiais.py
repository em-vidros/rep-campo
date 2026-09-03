#!/usr/bin/env python3
"""Gera rotas_oficiais.py a partir da planilha oficial de cidades e rotas.

    python3 scripts/gerar_rotas_oficiais.py "/caminho/CIDADES, ROTAS E TABELAS ...xlsx"

A planilha e a fonte da verdade. Rodar de novo sempre que ela for atualizada -
nunca editar rep_campo/dominio/rotas_oficiais.py na mao.
"""
import os
import sys
from datetime import date

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

import openpyxl

from rep_campo.dominio.texto import chave_cidade as chave


def main():
    if len(sys.argv) < 2:
        print(__doc__)
        return 1
    caminho = sys.argv[1]
    wb = openpyxl.load_workbook(caminho, data_only=True)
    ws = wb[wb.sheetnames[0]]

    cidades = []
    for linha in ws.iter_rows(min_row=2, values_only=True):
        if not linha or not linha[0]:
            continue
        cidade = str(linha[0]).strip()
        base = str(linha[1] or "").strip()
        rota = str(linha[2] or "").strip()
        tabela = str(linha[3] or "").strip()
        cidades.append((cidade, base, rota, tabela))

    # Migracao de 01/09/2026 decidida pelo Ricardo: a planilha de 28/08 ainda
    # nao refletia Alto Alegre do Pindare. Sobrepoe.
    MIGRADOS = {"alto alegre do pindare"}
    itz = [c for c in cidades
           if c[1].lower() == "imperatriz" or chave(c[0]) in MIGRADOS]
    rap = [c for c in cidades
           if c[1].lower() == "raposa" and chave(c[0]) not in MIGRADOS]

    def bloco(lista):
        # a planilha repete cidade (ex.: Santa Luzia do Pará em duas grafias
        # que normalizam igual). Última vale, como num dict literal.
        unicas = {}
        for cidade, _base, rota, tabela in sorted(lista):
            rota_limpa = rota if rota and chave(rota) != "sem rota" else ""
            unicas[chave(cidade)] = (rota_limpa, tabela)
        return "\n".join(
            '    %-42s (%-14s %s),' % ('"%s":' % k, '"%s",' % v[0], '"%s"' % v[1])
            for k, v in unicas.items())

    saida = '''# -*- coding: utf-8 -*-
"""Cidade -> rota e tabela de preco. GERADO AUTOMATICAMENTE, nao editar na mao.

Fonte: %s
Gerado em: %s por gerar_rotas_oficiais.py

A rota vazia significa "Sem Rota" na planilha: a cidade e atendida, mas nao
entra em roteiro regular. E diferente de cidade ausente da planilha, que nao
tem atendimento definido.
"""

# cidade normalizada -> (rota, tabela de preco)
CIDADES_ITZ = {
%s
}

CIDADES_RAPOSA = {
%s
}


EXPANSOES = [
    (r"\\bdo ma\\b", "do maranhao"), (r"\\bdo to\\b", "do tocantins"),
    (r"\\bdo pa\\b", "do para"), (r"\\bdo pi\\b", "do piaui"),
    (r"\\bgov\\.?\\b", "governador"), (r"\\bpres\\.?\\b", "presidente"),
    (r"\\bsto\\.?\\b", "santo"), (r"\\bsta\\.?\\b", "santa"),
]
CORRECOES = {"araguaiina": "araguaina"}


def _chave(cidade):
    import re
    import unicodedata
    t = unicodedata.normalize("NFKD", str(cidade or "")).encode("ascii", "ignore").decode()
    t = t.split("/")[0]
    t = re.sub(r"[-_.]+", " ", t).lower()
    t = re.sub(r"\\s+", " ", t).strip()
    for padrao, troca in EXPANSOES:
        t = re.sub(padrao, troca, t)
    t = re.sub(r"\\s+", " ", t).strip()
    return CORRECOES.get(t, t)


def rota_da_cidade(cidade):
    """Rota oficial da cidade na base Itz. None se nao estiver na planilha."""
    achado = CIDADES_ITZ.get(_chave(cidade))
    return achado[0] if achado else None


def tabela_da_cidade(cidade):
    achado = CIDADES_ITZ.get(_chave(cidade))
    return achado[1] if achado else None


def e_da_base_itz(cidade):
    """True se a planilha diz que a cidade e atendida pela base Imperatriz."""
    return _chave(cidade) in CIDADES_ITZ


def e_da_raposa(cidade):
    return _chave(cidade) in CIDADES_RAPOSA


ROTAS = sorted({r for r, _ in CIDADES_ITZ.values() if r})
TOTAL_ITZ = len(CIDADES_ITZ)
TOTAL_RAPOSA = len(CIDADES_RAPOSA)
''' % (caminho.split("/")[-1], date.today().strftime("%d/%m/%Y"), bloco(itz), bloco(rap))

    destino = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))),
                         "rep_campo", "dominio", "rotas_oficiais.py")
    with open(destino, "w", encoding="utf-8") as fh:
        fh.write(saida)

    print("[OK] %s gerado" % destino)
    print("  base Imperatriz: %d cidade(s)" % len(itz))
    print("  base Raposa    : %d cidade(s)" % len(rap))
    rotas = sorted({r for _, _, r, _ in itz if r and chave(r) != "sem rota"})
    print("  rotas da Itz   : %s" % ", ".join(rotas))
    return 0


if __name__ == "__main__":
    sys.exit(main())
