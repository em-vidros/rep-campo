# -*- coding: utf-8 -*-
"""Importação da planilha oficial de rotas pela tela (só admin). Borda fina."""
from io import BytesIO

from flask import Blueprint, jsonify, render_template, request, session

from rep_campo.dominio import catalogos as C
from rep_campo.infra import db as dbmod
from rep_campo.infra import repositorios as repo
from rep_campo.infra.relogio import agora
from rep_campo.web.acesso import admin_obrigatorio

bp = Blueprint("importar", __name__)


@bp.route("/importar")
@admin_obrigatorio
def pagina():
    return render_template("importar.html", nome=session.get("nome"),
                           papel=session.get("papel"))


@bp.route("/api/importar/rotas", methods=["POST"])
@admin_obrigatorio
def receber_planilha():
    arq = request.files.get("arquivo")
    if not arq or not arq.filename:
        return jsonify({"erro": "arquivo_ausente"}), 400
    dados = arq.read(C.MAX_ARQUIVO_IMPORTACAO + 1)
    if len(dados) > C.MAX_ARQUIVO_IMPORTACAO:
        return jsonify({"erro": "arquivo_grande"}), 413
    try:
        import openpyxl
        wb = openpyxl.load_workbook(BytesIO(dados), data_only=True, read_only=True)
    except Exception as exc:
        return jsonify({"erro": "nao_e_planilha", "detalhe": str(exc)[:160]}), 400

    ws = wb[wb.sheetnames[0]]
    linhas, cabecalho = [], None
    for linha in ws.iter_rows(values_only=True):
        if not linha or not linha[0]:
            continue
        if cabecalho is None:
            cabecalho = [str(c or "").strip().upper() for c in linha]
            if "CIDADE" not in cabecalho:
                return jsonify({"erro": "sem_coluna_cidade",
                                "colunas": cabecalho[:6]}), 400
            continue
        linhas.append([str(c or "").strip() for c in linha[:4]])
    if not linhas:
        return jsonify({"erro": "planilha_vazia"}), 400

    itz, rotas = repo.substituir_rotas(dbmod.get_db(), linhas, agora())
    return jsonify({"ok": True, "cidades": len(linhas), "base_itz": itz,
                    "rotas": rotas, "arquivo": arq.filename})


@bp.route("/api/importar/aplicar-rotas", methods=["POST"])
@admin_obrigatorio
def aplicar():
    db = dbmod.get_db()
    mapa = repo.mapa_rotas(db)
    if not mapa:
        return jsonify({"erro": "importe_a_planilha_antes"}), 400
    corrigidos, resumo, fora, outra_base = repo.aplicar_mapa_rotas(db, mapa)
    return jsonify({
        "ok": True, "corrigidos": len(corrigidos),
        "resumo": [{"de": d, "para": p, "clientes": n}
                   for (d, p), n in resumo.most_common()],
        "fora_da_planilha": sorted(set(fora))[:60],
        "total_fora": len(set(fora)),
        "de_outra_base": sorted(set(outra_base)),
    })
